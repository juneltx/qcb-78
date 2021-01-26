'''接口自动化测试的基本流程：
1、根据接口文档拟写接口测试用例，然后写在excel里
2、执行测试--利用requests库的requests函数执行，将执行结果获取
3、执行结果和预期结果进行对比，判断是否通过，得到最终结果
4、最终结果写入excel表格'''
#为了实现自动化，把每个接口封装成函数，以下进行实例
import requests,pprint
import jsonpath
#发送接口请求的函数
def api_func(url_api,data_api):
    head={"X-lemonban-Media-Type":"lemonban.v2"}
    response=requests.post(url=url_api,json=data_api,headers=head)
    return response.json()

#调用主体
url_api="http://8.129.91.152:8766/futureloan/member/register"
data_api={
  "mobile_phone":"15822222222",
  "pwd":12345678,
  "type":"1",
  "reg_name":"mengmeng"
}
result=api_func(url_api,data_api)
print(result)


#读取数据：第三方库，openpyxl----excel表格数据操作，
# 先在cmd中 pip install下，然后导入
#1、文档---加载进来---加载工作簿对象，load workbooks，可以把表格文件和代码复制在同级目录下，
#2、表单对象：
#3、获取单元格对象
#4、写入重新赋值，保存文件才能生效
from openpyxl import load_workbook#导入指定部分
# 1读取数据的函数
def read_data(filename,sheetname)
    wb=load_workbook(filename)#"test_case_api.xlsx"
    sheet=wb[sheetname]#读取表单对象
    cases=[]
    #获取最大行号
    max_row=sheet.max_row
    for i in range(2,max_row+1):
        case=dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,
        data=sheet.cell(row=i,column=6).value,
        expected_result=sheet.cell(row=i,column=7).value)
        cases.append(case)
    print(cases)
    return cases
#获取单元格值，然后重新赋值，即相当于写入
# print(cell.value)
# cell.value="用例编号"
# wb.save("test_case_api.xlsx")#保存文本，则修改了excel文件名字一样，为保存，不一样，则是另存为。
result=read_data("test_case_api.xlsx","login")


#2写入结果数据的函数
def write_result(filename,sheetname,final_result,row,column):
    wb=load_workbook(filename)
    sheet=wb[sheetname]
    sheet.cell(row=row,column=column).value=final_result
    wb.save(filename)

#3发送接口请求的函数
def api_func(url_api,data_api):
    head={"X-lemonban-Media-Type":"lemonban.v2"}
    response=requests.post(url=url_api,json=data_api,headers=head)
    return response.json()

#内容梳理
''''''